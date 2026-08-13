# -*- coding: utf-8 -*-
"""GK2A/AMI L1B DN -> 밝기온도(K) 변환.

2024_AWS_GK2A_RADAR/자료/gk2a_cal.py 를 가져오고 **IR105 를 추가**했다.
(원본에는 IR105 계수가 없었다 — WV063/WV073/IR087/IR112/IR123 뿐)

계수 출처: 국가기상위성센터(NMSC)
  20191115_gk-2a ami calibration table_v3.1_ir133_srf_shift.xlsx
  시트 'coeff.& equation_WN' 19행 (IR10.5)
API 허브가 주는 L1B .nc 에는 gain/offset 이 안 들어 있어 여기 박아둔다.

식(엑셀 원문):
  Radiance = DN2Rad_Gain * Count + DN2Rad_Offset       [mW/m^2/sr/cm^-1]
  Te = (hc/k * (v*100)) / ln( (2hc^2) * (v*100)^3 / (R*1e-5) + 1 )
  Tb = c0 + c1*Te + c2*Te^2

함정: Gain 이 음수라 DN 이 클수록 차갑다.
      DN 이 Rad2DN_Offset(IR105 는 8152.50)을 넘으면 Radiance <= 0 이라 무효.
      13bit(0~8191) 안이어도 끝쪽 39개는 버려야 한다.

검증: `python gk2a_cal.py` 를 실행하면 NMSC 배포표('Calibration Table_WN' 시트,
      IR 10.5 의 Brightness Temperature 열)와 직접 대조한다. openpyxl 필요.
"""
import numpy as np

# 물리상수 — 엑셀에 적힌 값 그대로. 다른 값 쓰면 소수점 아래가 어긋난다.
TWO_HC2 = 1.1910428681415875e-16      # 2hc^2
HC_K    = 1.4387769599838155e-2       # hc/k

# 채널별: (중심파수 cm^-1, DN2Rad_Gain, DN2Rad_Offset, c0, c1, c2, bits, Rad2DN_Offset)
COEF = {
    # ---- 추가분 (2026-08-11) ----
    # ★ SW038 은 **14bit** 다 — 다른 모든 채널이 12~13bit 라 lut() 표 크기부터 다르다.
    #   실측 파일 DN 범위 15,662~16,342 로 13bit(8191)를 훌쩍 넘는다.
    #   무효 문턱도 Rad2DN_Offset=16,344.00 이라 14bit 상한(16,383) 바로 아래다 -> 끝쪽 40개 무효.
    # ★ 주간 BT 는 태양반사가 섞인 **겉보기 온도**라 300K 를 넘는다(표상 DN 0 에서 400K).
    #   물리적 온도가 아니다. 정규화 범위를 IR105(180~330K)로 쓰면 주간이 통째로 클립된다.
    "SW038": (2612.67737352111, -1.08296517282724e-3,  17.699987411499,
              -0.447843939824124,     1.00065568090389,  -6.33824089912448e-8,   14, 16344.004272354),
    # -----------------------------
    "WV063": (1617.60924253134,  -1.08914673328399e-2,  44.177703857421797,
              -1.76279494011147,      1.00414910562278,  -9.8331091431938496e-7, 12, 4056.1755829),
    "WV073": (1365.24999202444,  -9.6982717514037999e-3, 79.060852050781193,
              -6.1312485969659498e-2, 1.0001900872294101, -1.0586365675049899e-7, 13, 8152.05575564),
    "IR087": (1164.94939285634,  -1.44806550815701e-2,  118.050903320312,
              -0.14141852820315501,   1.0005223290688501, -3.6287276076108999e-7, 13, 8152.31787894),
    # ---- 추가분 ----
    "IR105": (966.153383925781,  -1.98196955025196e-2,  161.580139160156,
              -0.142866448475266,     1.00064069572091,  -5.5044329495985e-7,    13, 8152.50361135),
    # ----------------
    "IR112": (891.71305730126005, -2.1674485877156199e-2, 176.71343994140599,
              -0.24911171849614799,   1.0012116687375601, -1.1316796401166499e-6, 13, 8153.06258903),
    "IR123": (810.60900787123001, -2.3379972204565998e-2, 190.64962768554599,
              -0.45811388572273798,   1.0024552097553501, -2.5306431472047599e-6, 13, 8154.39924468),
}


def dn_to_bt(dn, channel="IR105"):
    """DN(정수배열) -> 밝기온도 K. 무효화소는 NaN."""
    wn, gain, off, c0, c1, c2, _, _ = COEF[channel]
    rad = gain * np.asarray(dn, np.float64) + off       # mW/m^2/sr/cm^-1
    with np.errstate(divide="ignore", invalid="ignore"):
        rad_si = rad * 1e-5                              # -> W/m^2/sr/m^-1 (SI)
        v = wn * 100.0                                   # cm^-1 -> m^-1
        te = (HC_K * v) / np.log(TWO_HC2 * v**3 / rad_si + 1.0)
        tb = c0 + c1 * te + c2 * te * te
    return np.where(rad > 0, tb, np.nan)


def lut(channel="IR105"):
    """DN 을 인덱스로 바로 BT 를 뽑는 표(길이 2**bits).
    900x900 을 매 장 계산하는 것보다 훨씬 빠르다 — 캐시 빌드의 핵심."""
    bits = COEF[channel][6]
    return dn_to_bt(np.arange(1 << bits), channel).astype(np.float32)


# --------------------------------------------------------------------------- 검증
def _verify_against_nmsc_table(channel="IR105", xlsx=None, tol=1e-3):
    """NMSC 배포표의 Brightness Temperature 열과 직접 대조."""
    from openpyxl import load_workbook
    if xlsx is None:
        xlsx = (r"C:\Users\12345\Downloads\apcc\2024_AWS_GK2A_RADAR"
                r"\20191115_gk-2a ami calibration table_v3.1_ir133_srf_shift.xlsx")
    # 'Calibration Table_WN' 열 배치: 채널마다 (Radiance, Effective BT, BT) 3열.
    # col 0 = Count, VIS/NIR 6채널이 2열씩(1~12), 그 뒤 적외가 3열씩.
    #   IR3.8 -> 13,14,15 / IR6.3 -> 16,17,18 / ... / IR10.5 -> 31,32,33
    col_bt = {"SW038": 15, "IR105": 33, "IR112": 36}[channel]

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["Calibration Table_WN"]
    it = ws.iter_rows(values_only=True)
    next(it); next(it)                       # 헤더 2줄
    ref, cnt = [], []
    for row in it:
        if row[0] is None:
            continue
        cnt.append(int(row[0]))
        ref.append(np.nan if row[col_bt] is None else float(row[col_bt]))
    wb.close()

    cnt = np.asarray(cnt); ref = np.asarray(ref)
    m = cnt < (1 << COEF[channel][6])        # 13bit 범위만
    cnt, ref = cnt[m], ref[m]
    mine = dn_to_bt(cnt, channel)

    ok = np.isfinite(mine) & np.isfinite(ref)
    diff = np.abs(mine[ok] - ref[ok])
    print(f"[{channel}] 표와 대조: 비교 {ok.sum()}개 / 전체 {len(cnt)}개")
    print(f"          최대오차 {diff.max():.3e} K, 평균 {diff.mean():.3e} K")
    print(f"          -> {'통과' if diff.max() < tol else '★불일치★'}")
    return diff.max() < tol


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    for ch in COEF:
        t = lut(ch)
        ok = np.isfinite(t)
        print(f"{ch}: 유효 DN {ok.sum():5d}개, BT {np.nanmin(t):6.1f} ~ {np.nanmax(t):6.1f} K")

    print()
    for dn in (2569, 3748, 5146, 6430, 7735):      # 실제 파일 분위수
        print(f"  IR105  DN {dn:5d} -> {float(dn_to_bt(dn,'IR105')):6.2f} K")

    print()
    try:
        _verify_against_nmsc_table("IR105")
        _verify_against_nmsc_table("SW038")
    except ImportError:
        print("openpyxl 없음 -> 표 대조 건너뜀 "
              "(base 아나콘다로: C:\\ProgramData\\anaconda3\\python.exe gk2a_cal.py)")
